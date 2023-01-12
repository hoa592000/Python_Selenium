from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl as O

Execl_file = "Book1.xlsx"
Excel_worksheet= "Sheet1"
wb = O.load_workbook(Execl_file)
ws = wb[Excel_worksheet]
# row_num = ws.max_row
# col_num = ws.max_column

PATH = "chromedriver.exe"
driver = webdriver.Chrome(PATH)
driver.maximize_window()
# driver = webdriver.Chrome(PATH, chrome_options= chrome_option)
sleep(1)
driver.get("https://demoqa.com/text-box")
sleep(5)
textName = driver.find_element(By.ID, 'userName')
# textName.send_keys("Test1")
textEmail = driver.find_element(By.ID, 'userEmail')
# textEmail.send_keys("Test2")
textCurrentAddrest = driver.find_element(By.ID, 'currentAddress')
# textCurrentAddrest.send_keys("Test3")
textPermanetAddres = driver.find_element(By.ID, 'permanentAddress')
# textPermanetAddres.send_keys("Test4")
# btnSubmit = driver.find_element(By.CSS_SELECTOR, 'button[id="submit"][class="btn btn-primary"]')
# btnSubmit.click()
def intputText():
    for r in range(2, ws.max_row + 1):
        fullName = str(ws.cell(r,1).value)
        textName.send_keys(fullName)
        email = str(ws.cell(r,2).value)
        textEmail.send_keys(email)
        currentAddrest = str(ws.cell(r,3).value)
        textCurrentAddrest.send_keys(currentAddrest)
        permanetAddres = str(ws.cell(r,4).value)
        textPermanetAddres.send_keys(permanetAddres)
        driver.find_element(By.CSS_SELECTOR, 'button[id="submit"][class="btn btn-primary"]').click()
        # # btnSubmit.click()
        # textName.clear()
        # textEmail.clear()
        # textCurrentAddrest.clear()
        # textPermanetAddres.clear()
        sleep(15)
        # Result = str()
        # if str(Result)   
intputText()