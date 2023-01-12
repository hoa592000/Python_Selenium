import openpyxl as O
Execl_file = "Book1.xlsx"
Excel_worksheet= "Sheet1"
wb = O.load_workbook(Execl_file)
ws = wb[Excel_worksheet]
row_num = ws.max_row
col_num = ws.max_column
print("Row is ", row_num, "Colume", col_num)
row = 1
print("Value is ",ws.cell(row,3).value)