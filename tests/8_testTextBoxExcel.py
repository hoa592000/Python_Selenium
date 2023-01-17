from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as O

Execl_file = "Book1.xlsx"
Excel_worksheet= "Sheet1"
wb = O.load_workbook(Execl_file)
ws = wb[Excel_worksheet]
# row_num = ws.max_row
# col_num = ws.max_column

PATH = "chromedriver.exe"
driver = webdriver.Chrome(PATH)
driver.maximize_window()
# driver = webdriver.Chrome(PATH, chrome_options= chrome_option)
sleep(1)
driver.get("https://demoqa.com/text-box")
sleep(5)
textName = driver.find_element(By.ID, 'userName')
textEmail = driver.find_element(By.ID, 'userEmail')
textCurrentAddrest = driver.find_element(By.ID, 'currentAddress')
textPermanetAddres = driver.find_element(By.ID, 'permanentAddress')
sleep(2)
btnSubmit = driver.find_element(By.CSS_SELECTOR, 'button[id="submit"]')
# WebDriverWait wait
# btnSubmit.click()
def intputText():
    for r in range(2, ws.max_row + 1):
        fullName = str(ws.cell(r,1).value)
        email = str(ws.cell(r,2).value)
        currentAddrest = str(ws.cell(r,3).value)
        permanetAddres = str(ws.cell(r,4).value)
        if(fullName is None):
            print("Full Name invalidate")
        else:
            textName.send_keys(fullName)
        if(len(email)==0):
            print("Email invalidate")
        else:
            textEmail.send_keys(email)
        if(len(currentAddrest)==0):
            print("CurentAddrest invalidate")
        else:
            textCurrentAddrest.send_keys(currentAddrest)
        if(len(permanetAddres)==0):
            print("PermanetAddres invalidate")
        else:
            textPermanetAddres.send_keys(permanetAddres)
        # element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[id="submit"][class="btn.btn-primary"]')))
        # element.click()
        sleep(5)
        driver.find_element(By.CSS_SELECTOR, 'button[id="submit"][type="button"]').click()
        textName.clear()
        textEmail.clear()
        textCurrentAddrest.clear()
        textPermanetAddres.clear()
        sleep(15)
        # Result = str()
        # if str(Result)   
intputText()